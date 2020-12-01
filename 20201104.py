import pymysql
import pandas as pd
from openpyxl import load_workbook

# wb = load_workbook('城市省份.xlsx')
# ws = wb['Sheet1']
# maxrow=ws.max_row
# maxcol=ws.max_column
# print(maxrow,maxcol)
# for row in ws.rows:
#     for cell in row:
#         if cell != None:
#             print(cell.value)


conn = pymysql.connect("192.168.1.118", user="root", password="root1234", db="daniel")
cursor = conn.cursor()
print(cursor)

# sql = ("INSERT INTO EMPLOYEE(FIRST_NAME,LAST_NAME,AGE,SEX,INCOME) VALUES(%S,%S,%S,%S,%S)")
# sql1=("SELECT * FROM EMPLOYEE")
# sql2=("DROP TABLE EMPLOYEE")
# cursor.execute(sql2)
# myresult=cursor.fetchall()
# for x in myresult:
#     print(x)

CreateSql = ("CREATE TABLE customers (name varchar(255), adress varchar(255))")
InsertSql = ("INSERT INTO customers VALUES (%s,%s)")
VAL = [
    ('John', 'Highway 21'),
    ('Peter', 'Lowstreet 27'),
    ('Amy', 'Apple st 652'),
    ('Hannah', 'Mountain 21'),
    ('Michael', 'Valley 345'),
    ('Sandy', 'Ocean blvd 2'),
    ('Betty', 'Green Grass 1'),
    ('Richard', 'Sky st 331'),
    ('Susan', 'One way 98'),
    ('Vicky', 'Yellow Garden 2'),
    ('Ben', 'Park Lane 38'),
    ('William', 'Central st 954'),
    ('Chuck', 'Main Road 989'),
    ('Viola', 'Sideway 1633'),
    ('Michelle', 'Blue Village')]

# for v in VAL:
#     cursor.execute(InsertSql, v)
# cursor.execute(CreateSql)
cursor.execute("SELECT * FROM customers")
conn.commit()
myreulst = cursor.fetchall()
# for i in myreulst:
#     print(i)


# dict={"jack":2,"mike":3}
# str1=str(dict)
# print(str1)
#
# import json
#
# dict1=json.loads("\"" + str1 + "\"")
# print(dict1)

name = []
adress = []
for j in myreulst:
    i = list(j)
    name.append(i[0])
    adress.append(i[1])

newdf=pd.DataFrame({'name':name,'adress':adress})
newdf.to_excel('test.xlsx')
print('OK')
# DictTest={'students':"jack","school":"mit"}
# print(str(DictTest))
# print(DictTest['students'])
# print(tuple(DictTest))
# print(list(DictTest.values()))

# The dictionary could convert to the List,Tuple and String,
# Except String remaining List and Tuple could not transform to the dictionary.
DictTest = {'NAME': 'JACK', 'HEIGHT': 188}
ListKey = list(DictTest)
ListValue = list(DictTest.values())
Transform= dict(zip(ListKey, ListValue))
print(Transform)