from openpyxl import load_workbook

import pymysql

conn = pymysql.connect("192.168.1.118", user="root", password="root1234", db="daniel")
cursor = conn.cursor()
print(cursor)

wb = load_workbook('margin.xlsx')
ws = wb['Sheet3']

wb1 = load_workbook('RB20201102.xlsx')
ws1 = wb1['Sheet1']

LIST1 = []
LIST2 = []

"""
Reading the process list

"""

for row in ws1.rows:
    GRcolumn = row[1]
    GoodRetail = GRcolumn.value
    LIST1.append(GoodRetail)
"""
Reading the ratio list 

"""
for row in ws.rows:
    prices = row[0]
    ratio = row[1]
    ComparePrices = prices.value
    LIST2.append(ComparePrices)

for i in LIST1:
    print(min(LIST2, key=lambda x: abs(x - i)))

"""
For find the near num with good retail

"""

# b2 = ws.cell(row=2, column=2)
# print(b2.value)
#
# c2=ws.cell(row=3,column=1)
# c3=ws.cell(row=4,column=1)
# print(c2.value+c3.value)
#
# c8=ws.cell(row=8,column=2)
# c10=ws.cell(row=10,column=2)
# r8=ws.cell(row=8,column=1)
# r10=ws.cell(row=10,column=1)
# x=(c8.value-((c8.value-c10.value)/(r10.value-r8.value)))
# print(x)
# ws.cell(row=9,column=2).value=x
# wb.save('margin.xlsx')
# print('ok')
